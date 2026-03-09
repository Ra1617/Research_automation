import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# Color scheme
HEADER_FILL = PatternFill("solid", start_color="1F4E79")   # Dark blue
HIGH_FILL   = PatternFill("solid", start_color="E2EFDA")   # Light green
LOW_FILL    = PatternFill("solid", start_color="FFF2CC")   # Light yellow
NONE_FILL   = PatternFill("solid", start_color="FCE4D6")   # Light red/orange

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Arial", size=9)
CONF_FONT   = Font(name="Arial", size=8, italic=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def _write_headers(ws, all_params: list):
    meta_cols = ["Company Name", "Timestamp"]
    param_cols = [p["key"] for p in all_params]
    conf_cols  = [f"{p['key']} [confidence]" for p in all_params]

    # Interleave: value col, confidence col per parameter
    headers = meta_cols
    for pk, ck in zip(param_cols, conf_cols):
        headers.append(pk)
        headers.append(ck)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    return headers


def write_to_excel(company_name: str, result: dict, all_params: list, filepath: str = "output/company_data.xlsx"):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        # Find next empty row
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Company Data"
        headers = _write_headers(ws, all_params)
        ws.freeze_panes = "C2"  # freeze meta columns
        next_row = 2

    # Write data row
    col = 1
    # Meta columns
    ws.cell(row=next_row, column=col, value=company_name).font = Font(name="Arial", bold=True, size=9)
    ws.cell(row=next_row, column=col).border = THIN_BORDER
    col += 1
    ws.cell(row=next_row, column=col, value=ts).font = CELL_FONT
    ws.cell(row=next_row, column=col).border = THIN_BORDER
    col += 1

    # Parameter columns (value + confidence pairs)
    for p in all_params:
        key = p["key"]
        entry = result.get(key, {"value": None, "confidence": "none"})
        value = entry.get("value")
        score = entry.get("score", 0)
        confidence = "high" if score >= 80 else "low" if score > 0 else "none"

        # Value cell
        val_cell = ws.cell(row=next_row, column=col, value=str(value) if value is not None else "")
        val_cell.font = CELL_FONT
        val_cell.alignment = Alignment(wrap_text=True, vertical="top")
        val_cell.border = THIN_BORDER
        if confidence == "high":
            val_cell.fill = HIGH_FILL
        elif confidence == "low":
            val_cell.fill = LOW_FILL
        else:
            val_cell.fill = NONE_FILL
        col += 1

        # Confidence cell
        conf_cell = ws.cell(row=next_row, column=col, value=confidence)
        conf_cell.font = CONF_FONT
        conf_cell.alignment = Alignment(horizontal="center", vertical="top")
        conf_cell.border = THIN_BORDER
        col += 1

    ws.row_dimensions[next_row].height = 60

    wb.save(filepath)
    print(f"\n Excel saved: {filepath}")
    print(f"   Row {next_row}: {company_name}")


def write_summary_sheet(wb_path: str, summary_data: dict):
    """Add/update a summary sheet with stats."""
    wb = openpyxl.load_workbook(wb_path)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for row, (k, v) in enumerate(summary_data.items(), 2):
        ws.cell(row=row, column=1, value=k).font = CELL_FONT
        ws.cell(row=row, column=2, value=str(v)).font = CELL_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    wb.save(wb_path)
